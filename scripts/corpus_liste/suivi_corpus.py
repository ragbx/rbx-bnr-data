from os.path import join

import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ref_date = "20260502"

ref = pd.read_csv(
    join("results", "ref", f"_ref_files_{ref_date}.csv.gz"),
    usecols=["corpus_code", "uuid", "size", "conservation_statut", "file_type", "publication_statut"],
    low_memory=False,
)


def repartition(s):
    """Répartition des valeurs sous la forme "valeur (n) ; valeur (n)"."""
    vc = s.fillna("NON RENSEIGNÉ").value_counts()
    return " ; ".join(f"{k} ({v})" for k, v in vc.items())


# chaînes de caractères marquant, dans conservation_statut, un traitement s3 réalisé
STATUTS_TRAITES_S3 = ("TRANSFERT_S3_OK", "CORBEILLE", "SUPPRIMER", "NE PAS GARDER")


def traitement_s3(s):
    """Part (en %) des fichiers dont conservation_statut contient l'une des chaînes."""
    traites = s.apply(lambda v: any(chaine in str(v) for chaine in STATUTS_TRAITES_S3))
    return round(100 * traites.sum() / len(s), 1)


c = (
    ref.groupby("corpus_code")
    .agg(
        fichiers=("uuid", "count"),
        volume_go=("size", lambda s: round(s.sum() / 1e9, 2)),
        conservation_statut=("conservation_statut", repartition),
        traitement_s3=("conservation_statut", traitement_s3),
        fichiers_publies=("publication_statut", lambda s: (s == "oui").sum()),
    )
    .reset_index()
)

# une colonne par valeur de file_type
ft = pd.crosstab(ref["corpus_code"], ref["file_type"]).reset_index()
c = c.merge(ft, on="corpus_code", how="left")

l = pd.read_excel(join("data", "corpus_liste", "bnr_corpus.xlsx"))
# les colonnes calculées sont remplacées, les colonnes descriptives conservées
l = l.drop(columns=c.columns.drop("corpus_code"), errors="ignore")
l = l.merge(c, on="corpus_code", how="outer")

# ligne total
total = {
    "corpus_code": "TOTAL",
    "fichiers": len(ref),
    "volume_go": round(ref["size"].sum() / 1e9, 2),
    "conservation_statut": repartition(ref["conservation_statut"]),
    "traitement_s3": traitement_s3(ref["conservation_statut"]),
    "fichiers_publies": (ref["publication_statut"] == "oui").sum(),
}
total.update(ref["file_type"].value_counts().to_dict())
l = pd.concat([l, pd.DataFrame([total])], ignore_index=True)

out_path = join("results", "corpus_liste", "suivi_corpus.xlsx")
l.to_excel(out_path, index=False)

# mise en forme
wb = load_workbook(out_path)
ws = wb.active
colonnes = [cell.value for cell in ws[1]]
derniere = ws.max_row

# en-tête : gras sur fond gris, ligne figée
for cell in ws[1]:
    cell.font = Font(bold=True)
    cell.fill = PatternFill("solid", fgColor="D9D9D9")
ws.freeze_panes = "A2"

# ligne TOTAL : gras, bordure supérieure
for cell in ws[derniere]:
    cell.font = Font(bold=True)
    cell.border = Border(top=Side(style="medium"))

# traitement_s3 : échelle de couleurs rouge -> jaune -> vert
col_ts3 = get_column_letter(colonnes.index("traitement_s3") + 1)
ws.conditional_formatting.add(
    f"{col_ts3}2:{col_ts3}{derniere}",
    ColorScaleRule(
        start_type="num", start_value=0, start_color="F8696B",
        mid_type="num", mid_value=50, mid_color="FFEB84",
        end_type="num", end_value=100, end_color="63BE7B",
    ),
)

# séparateur de milliers sur les décomptes et le volume
for nom in ["fichiers", "volume_go", "fichiers_publies"] + ft.columns.drop("corpus_code").tolist():
    col = get_column_letter(colonnes.index(nom) + 1)
    fmt = "#,##0.00" if nom == "volume_go" else "#,##0"
    for cell in ws[col][1:]:
        cell.number_format = fmt

# largeur des colonnes
for idx, nom in enumerate(colonnes, start=1):
    largeur = 60 if nom in ("corpus", "conservation_statut") else max(12, len(str(nom)) + 2)
    ws.column_dimensions[get_column_letter(idx)].width = largeur

wb.save(out_path)

from os.path import join

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

ref_date = "20260502"

ref = pd.read_csv(
    join("results", "ref", f"_ref_files_{ref_date}.csv.gz"),
    usecols=["path", "conservation_statut"],
    low_memory=False,
)

# chaînes de caractères marquant, dans conservation_statut, un traitement s3 réalisé
STATUTS_TRAITES_S3 = ("TRANSFERT_S3_OK", "CORBEILLE", "SUPPRIMER", "NE PAS GARDER")


def repartition(s):
    """Répartition des valeurs sous la forme "valeur (n) ; valeur (n)"."""
    vc = s.fillna("NON RENSEIGNÉ").value_counts()
    return " ; ".join(f"{k} ({v})" for k, v in vc.items())


# fichiers dont le traitement s3 n'est pas réalisé
ko = ref[~ref["conservation_statut"].apply(lambda v: any(chaine in str(v) for chaine in STATUTS_TRAITES_S3))]

repertoires = (
    ko.groupby("path")
    .agg(
        nb_fichiers=("conservation_statut", "size"),
        conservation_statut=("conservation_statut", repartition),
    )
    .reset_index()
    .rename(columns={"path": "répertoire"})
    .sort_values("répertoire")
)

out_path = join("results", "s3", "repertoires_traitement_s3_ko.xlsx")
repertoires.to_excel(out_path, index=False)

# mise en forme
wb = load_workbook(out_path)
ws = wb.active

# en-tête : gras sur fond gris, ligne figée
for cell in ws[1]:
    cell.font = Font(bold=True)
    cell.fill = PatternFill("solid", fgColor="D9D9D9")
ws.freeze_panes = "A2"

# séparateur de milliers sur nb_fichiers
for cell in ws["B"][1:]:
    cell.number_format = "#,##0"

# largeur des colonnes
for idx, largeur in enumerate([90, 12, 60], start=1):
    ws.column_dimensions[get_column_letter(idx)].width = largeur

wb.save(out_path)
